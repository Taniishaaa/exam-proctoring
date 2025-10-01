# teacher.py (updated)
from xmlrpc.server import SimpleXMLRPCServer
from socketserver import ThreadingMixIn
import datetime
import threading
import xmlrpc.client
from pathlib import Path
import logging

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    raise SystemExit("Please install openpyxl: pip install openpyxl")

# ---------------- LOGGING ----------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | TEACHER | %(levelname)s | %(message)s",
    datefmt="%H:%M:%S"
)
logger = logging.getLogger("teacher")

class ThreadingXMLRPCServer(ThreadingMixIn, SimpleXMLRPCServer):
    daemon_threads = True

# Sample student data
students = {
    "1": {"name": "Swaroop", "marks": 0, "flag": 0},
    "2": {"name": "Tanisha", "marks": 0, "flag": 0},
    "3": {"name": "Siddhesh", "marks": 0, "flag": 0},
    "4": {"name": "Ayush", "marks": 0, "flag": 0},
    "5": {"name": "Nidhi", "marks": 0, "flag": 0},
}

local_time = None
excel_path = Path("results.xlsx")
_write_lock = threading.Lock()
results_ready = False

# ---------------- FUNCTIONS ----------------

def input_time():
    global local_time
    s = input("[Teacher] Enter local time (HH-MM-SS): ")
    local_time = datetime.datetime.strptime(s, "%H-%M-%S")
    logger.info("[Teacher] Time set.")
    return True

def calculate_cv(server_time_str):
    global local_time
    server_time = datetime.datetime.strptime(server_time_str, "%H-%M-%S")
    cv = (local_time - server_time).total_seconds()
    proxy = xmlrpc.client.ServerProxy("http://127.0.0.1:9000/", allow_none=True)
    proxy.receive_cv("Teacher", cv)
    return True

def apply_adjustment(adj):
    global local_time
    if local_time is not None:
        local_time = local_time + datetime.timedelta(seconds=float(adj))
        logger.info(f"[Teacher] Adjusted local time: {local_time.strftime('%H-%M-%S')}")
    return True

def start_exam():
    logger.info("[Teacher] Received start_exam()")
    return True

def deduct_marks(roll, flag):
    if roll in students:
        students[roll]["flag"] = flag
        if flag == 1:
            students[roll]["marks"] = round(students[roll]["marks"] * 0.8, 2)
        elif flag == 2:
            students[roll]["marks"] = 0
    return True

def update_mcq_marks(roll, marks_percent, mcq_total):
    roll = str(roll)
    with _write_lock:
        final_marks = round((float(marks_percent) / 100) * float(mcq_total), 2)

        if roll not in students:
            students[roll] = {
                "name": f"Student{roll}",
                "marks": final_marks,
                "flag": 0,
            }
        else:
            students[roll]["marks"] = final_marks

        logger.info(f"[Teacher] Roll {roll}: {marks_percent}% of {mcq_total} → {final_marks}")

        # Update Excel
        try:
            if not excel_path.exists():
                wb = Workbook()
                ws = wb.active
                ws.append(["Roll", "Name", "Marks/MCQ", "ISA"])
                for r, info in students.items():
                    ws.append([
                        r,
                        info.get("name", f"Student{r}"),
                        info.get("marks", "NA"),
                        "NA",
                    ])
                wb.save(excel_path)
            else:
                wb = load_workbook(excel_path)
                ws = wb.active
                updated = False
                for row in ws.iter_rows(min_row=2):
                    if row and str(row[0].value) == str(roll):
                        ws.cell(row=row[0].row, column=3, value=final_marks)
                        updated = True
                        break
                if not updated:
                    ws.append([
                        roll,
                        students[roll].get("name", f"Student{roll}"),
                        final_marks,
                        "NA",
                    ])
                wb.save(excel_path)
        except Exception as e:
            logger.error("[Teacher] ERROR updating Excel: %s", e)

    global results_ready
    results_ready = True
    return True

def get_results():
    ret = []
    for r, info in students.items():
        ret.append((r, info.get("name"), info.get("marks")))
    return ret

def release_results():
    import pandas as pd
    from openpyxl import load_workbook
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(row)  # (Roll, Name, Marks/MCQ, ISA)
    except Exception as e:
        logger.error("[Teacher] ERROR reading Excel: %s", e)
        return False

    logger.info("[Teacher] Results released to students.")

    proxy = xmlrpc.client.ServerProxy("http://127.0.0.1:9000/", allow_none=True)
    proxy.announce_results(data)
    return True

# ---------------- phase_complete RPC ----------------
def phase_complete(phase_name: str):
    logger.info("\n" + "="*40)
    logger.info(f"PHASE COMPLETE: {phase_name}")
    logger.info("="*40 + "\n")
    return True

def run_teacher():
    server = ThreadingXMLRPCServer(("0.0.0.0", 9001), allow_none=True, logRequests=False)
    server.register_function(input_time, "input_time")
    server.register_function(calculate_cv, "calculate_cv")
    server.register_function(apply_adjustment, "apply_adjustment")
    server.register_function(start_exam, "start_exam")
    server.register_function(deduct_marks, "deduct_marks")
    server.register_function(get_results, "get_results")
    server.register_function(release_results, "release_results")
    server.register_function(update_mcq_marks, "update_mcq_marks")
    server.register_function(phase_complete, "phase_complete")
    logger.info("[Teacher] Running on port 9001...")
    server.serve_forever()

# ---------------- MAIN ----------------
if __name__ == "__main__":
    import threading
    import time

    # Start RPC server in background
    threading.Thread(target=run_teacher, daemon=True).start()

    # Wait until results are ready
    while not results_ready:
        time.sleep(1)

    # Teacher manual release loop
    while True:
        choice = input("[Teacher] Do you want to release results to students? (y/n/exit): ").strip().lower()
        if choice == "y":
            release_results()
        elif choice == "n":
            logger.info("[Teacher] Results not released yet.")
        elif choice == "exit":
            logger.info("[Teacher] Exiting teacher console.")
            break
        else:
            logger.info("[Teacher] Please enter y/n/exit.")
