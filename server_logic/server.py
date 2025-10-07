import time
import datetime
import threading
import random
import heapq
import logging
import json
from typing import Dict, Set, List, Tuple, Any
from pathlib import Path
from xmlrpc.server import SimpleXMLRPCServer
from socketserver import ThreadingMixIn
import xmlrpc.client
import http.client
from collections import deque

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    raise SystemExit("Please install openpyxl: pip install openpyxl")

# ---------------- CONFIG ----------------
SERVER_HOST = "0.0.0.0"
SERVER_PORT = 9000
TEACHER_HOST = "127.0.0.1"
TEACHER_PORT = 9001
CLIENT_HOST = "127.0.0.1"
CLIENT_PORT = 9002
BACKUP_HOST = "127.0.0.1"
BACKUP_PORT = 9003

RPC_TIMEOUT = 5.0

# Replication / chunking defaults
DEFAULT_REPLICATION_FACTOR = 3
# chunk map: chunk_id -> list of roll strings
DEFAULT_CHUNK_MAP = {
    "chunk1": ["1", "2", "3"],
    "chunk2": ["4", "5"]
}
METADATA_PATH = Path("replication_metadata.json")

# ---------------- LOGGING ----------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | SERVER | %(levelname)s | %(message)s",
    datefmt="%H:%M:%S"
)
logger = logging.getLogger("server")

# -------------- (helper for xmlrpc timeouts) --------------
class TimeoutTransport(xmlrpc.client.Transport):
    def __init__(self, timeout=RPC_TIMEOUT):
        super().__init__()
        self._timeout = timeout
    def make_connection(self, host):
        return http.client.HTTPConnection(host, timeout=self._timeout)

def new_proxy(url: str, timeout=RPC_TIMEOUT):
    try:
        return xmlrpc.client.ServerProxy(url, allow_none=True, transport=TimeoutTransport(timeout))
    except Exception:
        return xmlrpc.client.ServerProxy(url, allow_none=True)

teacher_proxy = new_proxy(f"http://{TEACHER_HOST}:{TEACHER_PORT}/")
client_proxy  = new_proxy(f"http://{CLIENT_HOST}:{CLIENT_PORT}/")
backup_proxy = new_proxy(f"http://{BACKUP_HOST}:{BACKUP_PORT}/")

class ThreadingXMLRPCServer(ThreadingMixIn, SimpleXMLRPCServer):
    daemon_threads = True

# ---------------- ORIGINAL STATE ----------------
student_flags: Dict[str, int] = {}
terminated_students: Set[str] = set()
roll_to_name = {"1":"Swaroop","2":"Tanisha","3":"Siddhesh","4":"Ayush","5":"Nidhi"}

# Time sync
local_time = None
cvs = {}
cvs_lock = threading.Lock()

def input_time():
    global local_time
    s = input("[Server] Enter current local server time (HH-MM-SS): ")
    local_time = datetime.datetime.strptime(s, "%H-%M-%S")
    logger.info(f"[Server] Local time set to {local_time.strftime('%H-%M-%S')}")
    return True

def get_time():
    return local_time.strftime("%H-%M-%S") if local_time else ""

def receive_cv(source, cv):
    with cvs_lock:
        cvs[source] = float(cv)
    logger.info(f"[{datetime.datetime.now()}] [Server] Received CV from {source}: {cv}")
    return True

# ---------------- MCQ additions  ----------------
EXAM_DURATION = 300  # 5 minutes (seconds) for MCQ/exam

# Hardcoded 10 MCQs 
MCQ_QUESTIONS = {
    1: {"q": "Which protocol is used for time synchronization in this project?",
        "options": ["Lamport", "Berkeley", "Ricart-Agrawala", "HTTP"], "answer": 2},
    2: {"q": "Which algorithm ensures mutual exclusion for ISA marks?",
        "options": ["Token Ring", "Ricart-Agrawala", "Paxos", "Bully"], "answer": 2},
    3: {"q": "Which library writes Excel files?",
        "options": ["pandas", "openpyxl", "xlrd", "xlsxwriter"], "answer": 2},
    4: {"q": "What is the exam duration now set to (seconds)?",
        "options": ["20", "60", "300", "600"], "answer": 3},
    5: {"q": "Which RPC mechanism is used between nodes?",
        "options": ["gRPC", "XML-RPC", "REST", "WebSocket"], "answer": 2},
    6: {"q": "A warning (first cheating) reduces MCQ marks to what percent?",
        "options": ["100%", "80%", "50%", "0%"], "answer": 2},
    7: {"q": "If a student receives 2 warnings, MCQ marks become:",
        "options": ["100%", "80%", "50%", "0%"], "answer": 4},
    8: {"q": "Total MCQ marks possible in this test are:",
        "options": ["50", "70", "100", "120"], "answer": 3},
    9: {"q": "Who coordinates registration of student peer URLs?",
        "options": ["Teacher", "Client", "Server", "Student"], "answer": 3},
    10: {"q": "Which data structure logs RA intents in server.py?",
         "options": ["list", "heap", "set", "dict"], "answer": 2}
}

# MCQ state
mcq_lock = threading.Lock()
mcq_active = False
mcq_start_time = None
mcq_deadline = None
# student-specific answers: student_roll -> {qnum:int}
mcq_student_answers: Dict[str, Dict[int,int]] = {}
mcq_submitted_students: Set[str] = set()
mcq_final_scores: Dict[str,int] = {}

# Student registry and RA state (preserved)
students_registry: Dict[str,str] = {}
students_lock = threading.Lock()
isa_completed: Set[str] = set()
isa_lock = threading.Lock()
RA_MODE = True
_intent_heap: List[Tuple[int,str]] = []
_intent_lock = threading.Lock()
excel_path = Path("results.xlsx")
isa_ok_counts: Dict[str, Set[str]] = {}

# replication metadata (master keeps this)
replication_metadata: Dict[str, Any] = {}
replication_lock = threading.Lock()

# ---------------- MCQ API  ----------------
def start_mcq():
    """
    Start MCQ session (called when exam starts).
    """
    global mcq_active, mcq_start_time, mcq_deadline, mcq_student_answers, mcq_submitted_students, mcq_final_scores
    with mcq_lock:
        mcq_active = True
        mcq_start_time = time.time()
        mcq_deadline = mcq_start_time + EXAM_DURATION
        mcq_student_answers = {}
        mcq_submitted_students = set()
        mcq_final_scores = {}
    logger.info(f"[{datetime.datetime.now()}] [Server] MCQ started for duration {EXAM_DURATION} seconds.")
    return True

def get_mcq_active():
    with mcq_lock:
        return mcq_active

def get_question_for_student(roll: str, qnum: int):
    """
    Return the question object for given qnum.
    qnum is 1..10
    """
    q = MCQ_QUESTIONS.get(int(qnum))
    if not q:
        return {}
    # return q text and options
    return {"qnum": int(qnum), "q": q["q"], "options": q["options"]}

def submit_mcq_answer(roll: str, qnum: int, answer):
    """
    Students call this to record an answer for a question.
    answer: 1..4 for options, 0 or None for skip.
    Each student proceeds at their own pace so server just stores the given answer.
    """
    roll = str(roll)
    qnum = int(qnum)
    try:
        ans_i = int(answer)
    except Exception:
        ans_i = 0
    with mcq_lock:
        mcq_student_answers.setdefault(roll, {})[qnum] = ans_i
    logger.info(f"[{datetime.datetime.now()}] [Server] Recorded answer roll={roll} q={qnum} ans={ans_i}")
    return True

# submission window (sliding) using ms timestamps
SUBMISSION_WINDOW_MS = 1000   # group submissions that happen within this many ms
CAP = 3                       # main server handles first CAP submissions in each window
submission_history = deque()  # elements are (ts_ms:int, roll:str)
submission_lock = threading.Lock()

def _finalize_and_record(roll: str):
    """Existing local logic separated for reuse."""
    roll = str(roll)
    with mcq_lock:
        if roll in mcq_submitted_students:
            return True
        answers = mcq_student_answers.get(roll, {})
        raw = 0
        for qnum, qdef in MCQ_QUESTIONS.items():
            given = int(answers.get(qnum, 0) or 0)
            if given != 0 and given == qdef["answer"]:
                raw += 10

        final = raw
        mcq_final_scores[roll] = final
        mcq_submitted_students.add(roll)

    #logger.info(f"[{datetime.datetime.now()}] [Server] Finalized MCQ roll={roll} raw={raw} flags={flags} final={final}")

    # notify teacher
    try:
        teacher_proxy.update_mcq_marks(str(roll), int(final))
    except Exception as e:
        logger.warning("[Server] WARN teacher.update_mcq_marks: %s", e)

    # update excel 
    try:
        with _intent_lock:
            if not excel_path.exists():
                wb = Workbook()
                ws = wb.active
                ws.append(["Roll","Name","Marks","MCQ","ISA"])
                wb.save(excel_path)
            wb = load_workbook(excel_path)
            ws = wb.active
            updated = False
            for row in ws.iter_rows(min_row=2):
                if row and str(row[0].value) == str(roll):
                    while ws.max_column < 5:
                        ws.cell(row=1, column=ws.max_column+1, value=None)
                    ws.cell(row=row[0].row, column=4, value=int(final))
                    updated = True
                    break
            if not updated:
                ws.append([roll, roll_to_name.get(roll, f"Student{roll}"), "NA", int(final), "NA"])
            wb.save(excel_path)
    except Exception as e:
        logger.warning("[Server] WARN writing MCQ result to excel: %s", e)

    return True

def submit_mcq_final(roll: str):
    """
    Student requests final submission.
    If more than CAP submissions occur within SUBMISSION_WINDOW_MS,
    overflow is redirected to backup to ensure 3/2 split.
    """
    roll = str(roll)
    now_ms = int(time.time() * 1000)

    with submission_lock:
        # purge old entries older than window
        cutoff = now_ms - SUBMISSION_WINDOW_MS
        while submission_history and submission_history[0][0] < cutoff:
            submission_history.popleft()

        # count current window size and append this submission
        submission_history.append((now_ms, roll))
        window_count = len(submission_history)
        logger.info(f"[{datetime.datetime.now()}] [Server] submit_mcq_final called for roll={roll} (window_count_before={window_count-1})")

        # Determine which submissions stay local vs go to backup
        local_limit = 3  # first 3 stay on main
        idx = list(submission_history).index((now_ms, roll))
        if idx >= local_limit:
            # redirect this submission to backup
            try:
                answers = mcq_student_answers.get(roll, {})
                # Convert keys to strings for XML-RPC
                answers_str_keys = {str(k): v for k, v in answers.items()}
                logger.info(f"[{datetime.datetime.now()}] [Server] Redirecting roll={roll} to backup {BACKUP_HOST}:{BACKUP_PORT}")
                backup_proxy.submit_mcq_final(roll, answers_str_keys)
                return True
            except Exception as e:
                logger.error(f"[{datetime.datetime.now()}] [Server] ERROR redirecting to backup: {e}. Falling back to local processing.")

    # process locally
    return _finalize_and_record(roll)


# API for backup to push computed results back to main server
def accept_backup_result(roll: str, final_mark: int):
    """
    Called by backup server to inform main server of a finalized mark for roll.
    Updates teacher, Excel, and unified state.
    """
    roll = str(roll)
    try:
        final = int(final_mark)
    except Exception:
        final = int(float(final_mark))

    with mcq_lock:
        if roll in mcq_submitted_students:
            logger.info(f"[{datetime.datetime.now()}] [Server] accept_backup_result: roll={roll} already submitted locally, ignoring.")
            return True

        mcq_final_scores[roll] = final
        mcq_submitted_students.add(roll)

    logger.info(f"[{datetime.datetime.now()}] [Server] accept_backup_result: recorded roll={roll} final={final} (from backup)")

    # Notify teacher with mcq_total
    total_marks = len(MCQ_QUESTIONS) * 10
    try:
        teacher_proxy.update_mcq_marks(str(roll), int(final), mcq_total=total_marks)
    except Exception as e:
        logger.warning("[Server] WARN teacher.update_mcq_marks (from backup): %s", e)

    # Update Excel (unchanged)
    try:
        with _intent_lock:
            if not excel_path.exists():
                wb = Workbook()
                ws = wb.active
                ws.append(["Roll","Name","Marks","MCQ","ISA"])
                wb.save(excel_path)
            wb = load_workbook(excel_path)
            ws = wb.active
            updated = False
            for row in ws.iter_rows(min_row=2):
                if row and str(row[0].value) == str(roll):
                    while ws.max_column < 4:
                        ws.cell(row=1, column=ws.max_column+1, value=None)
                    ws.cell(row=row[0].row, column=4, value=int(final))
                    updated = True
                    break
            if not updated:
                ws.append([roll, roll_to_name.get(roll, f"Student{roll}"), "NA", int(final), "NA"])
            wb.save(excel_path)
    except Exception as e:
        logger.warning("[Server] WARN writing MCQ result to excel (from backup): %s", e)

    return True

def _auto_submit_pending_mcq():
    """
    Auto-submit any students who haven't submitted by deadline or when exam completes.
    Ensures first 3 go to main server, remaining go to backup if necessary.
    """
    with mcq_lock:
        pending = [r for r in students_registry.keys()
                   if r not in mcq_submitted_students and r not in terminated_students]

    logger.info(f"[{datetime.datetime.now()}] [Server] Auto-submitting pending MCQ for: {pending}")

    for idx, r in enumerate(pending):
        try:
            if idx < 3:
                # process locally
                submit_mcq_final(r)
            else:
                # redirect to backup
                answers = mcq_student_answers.get(r, {})
                answers_str_keys = {str(k): v for k, v in answers.items()}
                try:
                    backup_proxy.submit_mcq_final(r, answers_str_keys)
                except Exception as e:
                    logger.error(f"[{datetime.datetime.now()}] [Server] ERROR redirecting auto-submission roll={r} to backup: {e}")
                    # fallback to local
                    submit_mcq_final(r)

            # notify student of auto-submission
            if r in students_registry:
                try:
                    new_proxy(students_registry[r]).notify_mcq_submitted()
                except Exception as e:
                    logger.warning(f"[Server] WARN notify_mcq_submitted to {r}: {e}")

        except Exception as e:
            logger.error(f"[{datetime.datetime.now()}] [Server] ERROR auto-submitting for {r}: {e}")


# ---------------- ORIGINAL function updated to call MCQ start  ----------------
def _do_synchronization_background():
    global local_time, cvs
    logger.info("[Server] Starting Berkeley sync (background)...")
    with cvs_lock:
        cvs = {}
    server_time_str = get_time()
    if not server_time_str:
        logger.warning("[Server] Local time not set; call input_time() first.")
        return
    try:
        teacher_proxy.calculate_cv(server_time_str)
    except Exception as e:
        logger.warning("[Server] WARN teacher.calculate_cv: %s", e)
    try:
        client_proxy.calculate_cv(server_time_str)
    except Exception as e:
        logger.warning("[Server] WARN client.calculate_cv: %s", e)

    with cvs_lock:
        cvs["Server"] = 0.0

    deadline = time.time() + 12.0
    while time.time() < deadline:
        with cvs_lock:
            if {"Server", "Teacher", "Client"}.issubset(set(cvs.keys())):
                break
        time.sleep(0.05)

    with cvs_lock:
        if not {"Server", "Teacher", "Client"}.issubset(set(cvs.keys())):
            logger.warning("[Server] WARNING: Not all CVs arrived; proceeding with available CVs.")
        avg_cv = sum(cvs.values()) / max(1, len(cvs))

    adjustments = {node: avg_cv - cv for node, cv in cvs.items()}

    try:
        teacher_proxy.apply_adjustment(adjustments.get("Teacher", 0.0))
    except Exception:
        pass
    try:
        client_proxy.apply_adjustment(adjustments.get("Client", 0.0))
    except Exception:
        pass

    if local_time is not None:
        delta = adjustments.get("Server", 0.0)
        local_time += datetime.timedelta(seconds=delta)
        logger.info(f"[Server] Adjusted local time by {delta}s -> {local_time.strftime('%H-%M-%S')}")

    # The exam is no longer started automatically.
    # Simply announce that the Time-Sync phase is complete;
    # the admin will explicitly run "start_exam" from the panel.
    logger.info("[Server] Time Synchronization phase complete. Waiting for admin to start the exam.")
    try:
        teacher_proxy.phase_complete("Time Synchronization")
    except Exception:
        pass
    try:
        client_proxy.phase_complete("Time Synchronization")
    except Exception:
        pass
    # You can also broadcast phase_complete to all students here if desired:
    try:
        with students_lock:
            for r, url in students_registry.items():
                try:
                    new_proxy(url).phase_complete("Time Synchronization")
                except Exception:
                    pass
    except Exception:
        pass


def start_synchronization():
    threading.Thread(target=_do_synchronization_background, daemon=True).start()
    return True
"""
def cheating_detection(roll):
    name = roll_to_name.get(str(roll), "Unknown")
    if str(roll) in terminated_students:
        return None

    student_flags.setdefault(str(roll), 0)
    student_flags[str(roll)] += 1
    flag = student_flags[str(roll)]

    try:
        teacher_proxy.deduct_marks(str(roll), flag)
    except Exception:
        pass

    if flag == 1:
        msg = f"{roll}, {name} warning "

    elif flag == 2:
        msg = f"{roll}, {name} exam terminated"
        terminated_students.add(str(roll))

        # --- Record termination in Excel ---
        try:
            with _intent_lock:
                if not excel_path.exists():
                    wb = Workbook()
                    ws = wb.active
                    ws.append(["Roll","Name","Marks","MCQ","ISA"])
                    wb.save(excel_path)
                wb = load_workbook(excel_path)
                ws = wb.active
                updated = False

                for row in ws.iter_rows(min_row=2):
                    if row and str(row[0].value) == str(roll):
                        ws.cell(row=row[0].row, column=4, value=0)  # MCQ = 0
                        ws.cell(row=row[0].row, column=3, value="Terminated")
                        updated = True
                        break

                if not updated:
                    # find correct insertion index by roll order
                    insert_at = ws.max_row + 1  # default append
                    for row in ws.iter_rows(min_row=2):
                        existing_roll = int(row[0].value)
                        if int(roll) < existing_roll:
                            insert_at = row[0].row
                            break

                    ws.insert_rows(insert_at)
                    ws.cell(row=insert_at, column=1, value=roll)
                    ws.cell(row=insert_at, column=2, value=name)
                    ws.cell(row=insert_at, column=3, value="Terminated")
                    ws.cell(row=insert_at, column=4, value=0)
                    ws.cell(row=insert_at, column=5, value="NA")

                wb.save(excel_path)
        except Exception as e:
            logger.warning(f"[Server] WARN writing terminated student {roll} to Excel: {e}")


        # --- Notify student of termination ---
        try:
            if str(roll) in students_registry:
                new_proxy(students_registry[str(roll)]).notify_exam_terminated()
        except Exception as e:
            logger.warning(f"[Server] WARN notify_exam_terminated to {roll}: {e}")

    else:
        msg = f"{roll}, {name} no action"

    logger.info(f"[Server] Cheating detected: {msg}")
    return msg
"""

# ISA orchestration / RA helpers
students_registry: Dict[str,str] = {}
students_lock = threading.Lock()
isa_completed: Set[str] = set()
isa_lock = threading.Lock()
RA_MODE = True
_intent_heap: List[Tuple[int,str]] = []
_intent_lock = threading.Lock()
excel_path = Path("results.xlsx")
isa_ok_counts: Dict[str, Set[str]] = {}

def register_student(roll: str, url: str):
    with students_lock:
        students_registry[str(roll)] = str(url)
    logger.info(f"[{datetime.datetime.now()}] [Server] Registered student {roll} -> {url}")
    return True

def get_registry():
    with students_lock:
        return dict(students_registry)

def register_intent(roll: str, ts: float):
    try:
        ts_i = int(float(ts))
    except Exception:
        ts_i = int(time.time() * 1000000)
    with _intent_lock:
        heapq.heappush(_intent_heap, (ts_i, str(int(roll))))
    logger.info(f"[{datetime.datetime.now()}] [Server] Intent registered: roll={roll}, ts={ts_i}")
    _print_intent_queue()
    return True

def _print_intent_queue():
    with _intent_lock:
        q = sorted(_intent_heap)
    readable = [(ts, str(r)) for ts, r in q]
    logger.info(f"[{datetime.datetime.now()}] [Server] Current intent queue: {readable}")

def ok_signal(from_roll: str, to_roll: str):
    with isa_lock:
        isa_ok_counts.setdefault(str(to_roll), set()).add(str(from_roll))
        needed = set(students_registry.keys()) - {str(to_roll)}
        got = isa_ok_counts[str(to_roll)]
    logger.info(f"[{datetime.datetime.now()}] [Server] OK from {from_roll} -> {to_roll} ({len(got)}/{len(needed)})")
    return True

def update_isa(roll: str, isa_value: int):
    with isa_lock:
        tries = 5
        saved = False
        for i in range(tries):
            try:
                if not excel_path.exists():
                    wb = Workbook()
                    ws = wb.active
                    ws.append(["Roll","Name","Marks","MCQ","ISA"])
                    wb.save(excel_path)
                wb = load_workbook(excel_path)
                ws = wb.active
                updated = False
                for row in ws.iter_rows(min_row=2):
                    if row and str(row[0].value) == str(roll):
                        while ws.max_column < 4:
                            ws.cell(row=1, column=ws.max_column+1, value=None)
                        ws.cell(row=row[0].row, column=5, value=int(isa_value))
                        updated = True
                        break
                if not updated:
                    ws.append([roll, roll_to_name.get(roll, f"Student{roll}"), "NA", "NA", int(isa_value)])
                wb.save(excel_path)
                saved = True
                _sync_replicas_from_master()
                break
            except PermissionError:
                logger.warning("[Server] Excel file locked. Retrying...")
                time.sleep(0.5)
            except Exception as e:
                logger.error("[Server] Unexpected error saving excel: %s", e)
                break
        if not saved:
            logger.error("[Server] Failed to save ISA after retries.")

        isa_completed.add(str(roll))
        logger.info(f"[{datetime.datetime.now()}] [Server] Roll {roll} updated ISA={isa_value} and exited CS.")

    with _intent_lock:
        new_heap = [(ts, r) for ts, r in _intent_heap if str(r) != str(int(roll))]
        heapq.heapify(new_heap)
        _intent_heap.clear()
        _intent_heap.extend(new_heap)

    _print_intent_queue()
    _check_done_ra()
    return True

def _maybe_start_isa():
    global isa_phase
    with isa_lock:
        if isa_phase:
            return
        if not students_registry:
            logger.info("[Server] No students registered yet. Waiting...")
            return
        isa_phase = True
    pending = [r for r in students_registry.keys() if r not in isa_completed]
    if not pending:
        logger.info("[Server] No pending students for ISA.")
        return
    first = random.choice(pending)
    logger.info(f"[Server] Randomly selected first student: roll {first}")
    _announce_selection(first)

def _announce_selection(target_roll: str):
    isa_ok_counts[target_roll] = set()
    for roll, url in students_registry.items():
        try:
            new_proxy(url).notify_selection(target_roll)
        except Exception as e:
            logger.warning(f"[Server] WARN notify_selection failed for {roll}: {e}")
    logger.info(f"[{datetime.datetime.now()}] [Server] Selection broadcast complete for target {target_roll}")

def _proceed_next(just_finished_roll: str):
    pending = [r for r in students_registry.keys() if r not in isa_completed]
    if not pending:
        logger.info("[Server] ISA phase completed for all registered students (legacy).")
        for roll, url in students_registry.items():
            try:
                new_proxy(url).isa_phase_done(str(excel_path.resolve()))
            except Exception:
                pass
        return
    next_roll = random.choice(pending)
    logger.info(f"[Server] Next randomly selected student: roll {next_roll}")
    _announce_selection(next_roll)

def _check_done_ra():
    pending = [r for r in students_registry.keys() if r not in isa_completed]
    if not pending:
        logger.info(f"[{datetime.datetime.now()}] [Server] (RA_MODE) ISA phase completed for all students.")
        for roll, url in students_registry.items():
            try:
                new_proxy(url).isa_phase_done(str(excel_path.resolve()))
            except Exception:
                pass
        # At this point ISA is DONE; prompt admin to start replication
        #threading.Thread(target=_prompt_and_create_replication, daemon=True).start()

def exam_completed():
    """
    Called by trigger.py when exam ends. We must:
      - Ensure MCQ auto-submission for any pending students.
      - Fetch teacher results and write initial results to Excel.
      - Broadcast ask_to_request to students (starting ISA phase).
    """
    logger.info("[Server] Exam completion received. Finalizing MCQ submissions...")
    # Ensure MCQ auto-submit
    _auto_submit_pending_mcq()

    # Fetch results from Teacher & write initial results to Excel
    try:
        results = teacher_proxy.get_results()
        if not excel_path.exists():
            wb = Workbook()
            ws = wb.active
            ws.append(["Roll","Name","Marks","MCQ","ISA"])
            for row in results:
                if len(row) == 3:
                    r, name, marks = row
                    mcq_val = "NA"
                elif len(row) >= 4:
                    r, name, marks, mcq_val = row[0], row[1], row[2], row[3]
                else:
                    r, name, marks = row[0], row[1], row[2]
                    mcq_val = "NA"
                ws.append([r, name, marks, mcq_val, "NA"])
            wb.save(excel_path)
        logger.info(f"[Server] Wrote initial results to {excel_path.resolve()}")
    except Exception as e:
        logger.error(f"[Server] ERROR fetching results from Teacher: {e}")

    # Now broadcast ask_to_request to all students (start ISA phase)
    logger.info("[Server] Broadcasting ask_to_request to all students...")
    with students_lock:
        for roll, url in students_registry.items():
            try:
                new_proxy(url).ask_to_request()
                logger.info(f"[Server] ask_to_request called on roll {roll}")
            except Exception as e:
                logger.warning(f"[Server] WARN ask_to_request failed for {roll}: {e}")
    return True


def announce_results(data):
    # Called by teacher, forward results to all students
    for roll, url in students_registry.items():
        try:
            p = xmlrpc.client.ServerProxy(url, allow_none=True)
            p.show_results(data)
        except Exception as e:
            logger.error(f"[Server] ERROR sending results to {roll}: {e}")
    return True


# ---------------- Replication & Chunking (HDFS-like) ----------------
def _read_results_rows():
    """
    Read results.xlsx and return header + rows (each row is tuple of values)
    """
    if not excel_path.exists():
        logger.error("[Server] Results file not found: results.xlsx")
        return None, []
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        header = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(tuple(row))
        return header, rows
    except Exception as e:
        logger.error(f"[Server] Error reading results.xlsx: {e}")
        return None, []

def _filter_rows_for_rolls(rows, rolls: List[str]):
    """
    rows: list of tuples where first column is roll
    returns subset of rows matching rolls (string/int tolerant)
    """
    rolls_set = set(str(r) for r in rolls)
    selected = [r for r in rows if str(r[0]) in rolls_set]
    return selected

def _write_chunk_excel(filepath: Path, header: List[str], rows: List[Tuple]):
    """
    Write header + rows to an Excel file at filepath
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.append(header)
        for r in rows:
            ws.append(list(r))
        wb.save(filepath)
        logger.info(f"[Server] Wrote chunk file: {filepath}")
        return True
    except Exception as e:
        logger.error(f"[Server] Failed writing chunk file {filepath}: {e}")
        return False

def create_replicas_and_chunks(replication_factor: int = DEFAULT_REPLICATION_FACTOR, chunk_map: Dict[str, List[str]] = DEFAULT_CHUNK_MAP):
    """
    Create replicas and chunks from the authoritative results.xlsx.
    - replication_factor: number of replicas to create
    - chunk_map: mapping chunk_id -> list of roll strings
    Persist metadata to replication_metadata.json and keep in memory.
    """
    header, rows = _read_results_rows()
    if header is None:
        logger.error("[Server] Cannot create replicas: results.xlsx unavailable or unreadable.")
        return False

    meta = {
        "created_at": datetime.datetime.now().isoformat(),
        "replication_factor": int(replication_factor),
        "chunks": {},  # chunk_id -> rolls
        "replicas": {}  # replica_id -> {chunk_id: filepath, ...}
    }

    # Prepare chunks content
    chunk_contents: Dict[str, List[Tuple]] = {}
    for chunk_id, rolls in chunk_map.items():
        selected = _filter_rows_for_rolls(rows, rolls)
        chunk_contents[chunk_id] = selected
        meta["chunks"][chunk_id] = {"rolls": list(rolls), "count": len(selected)}

    # Create replicas
    for r_idx in range(1, int(replication_factor)+1):
        replica_key = f"replica_{r_idx}"
        meta["replicas"][replica_key] = {}
        for chunk_id, content_rows in chunk_contents.items():
            filename = f"replica_{r_idx}_{chunk_id}.xlsx"
            path = Path(filename).resolve()
            ok = _write_chunk_excel(path, header, content_rows)
            if not ok:
                logger.error(f"[Server] Failed to create replica file {filename}")
            # store path and basic metadata
            meta["replicas"][replica_key][chunk_id] = {
                "path": str(path),
                "rows": len(content_rows)
            }

    # persist metadata
    try:
        with replication_lock:
            global replication_metadata
            replication_metadata = meta
            with open(METADATA_PATH, "w", encoding="utf-8") as fh:
                json.dump(meta, fh, indent=2)
        logger.info(f"[Server] Replication metadata written to {METADATA_PATH}")
    except Exception as e:
        logger.error(f"[Server] Failed writing replication metadata: {e}")
        return False


    global chunk_locks
    chunk_locks = {}
    for replica_id, chunks in replication_metadata["replicas"].items():
        for chunk_id in chunks.keys():
            key = f"{replica_id}:{chunk_id}"
            chunk_locks[key] = ChunkLock(key)
    logger.info(f"[Server] Completed creating {replication_factor} replicas and {len(chunk_map)} chunks per replica.")
    return True

def _prompt_and_create_replication():
    """
    Called when ISA completes; prompts admin on server terminal whether to create replicas & chunks.
    Runs in a daemon thread to avoid blocking RA flows.
    """
    # small delay so the 'ISA completed' logs appear clearly first
    time.sleep(0.2)
    try:
        # Prompt
        ans = input("\n[Server] ISA finished. Start replication and chunk creation? (y/n): ").strip().lower()
        if ans and ans[0] == "y":
            logger.info("[Server] Admin accepted replication. Creating replicas & chunks...")
            ok = create_replicas_and_chunks()
            if ok:
                logger.info("[Server] Replication & chunk creation successful.")
            else:
                logger.error("[Server] Replication & chunk creation failed. See logs.")
        else:
            logger.info("[Server] Admin declined replication at this time.")
    except Exception as e:
        logger.error(f"[Server] Error during replication prompt/creation: {e}")
        
 
 
def _sync_replicas_from_master():
    """Refresh replica chunk files from results.xlsx after any update."""
    header, rows = _read_results_rows()
    if header is None:
        return
    with replication_lock:
        for replica_id, chunks in replication_metadata.get("replicas", {}).items():
            for chunk_id, info in chunks.items():
                path = Path(info["path"])
                rolls = DEFAULT_CHUNK_MAP.get(chunk_id, [])
                selected = _filter_rows_for_rolls(rows, rolls)
                _write_chunk_excel(path, header, selected)
                
    
# ---------------- Consistency & Lock Manager ----------------  
    
class ChunkLock:
    """Readers-writers lock with writer-preference for a single chunk id."""
    def __init__(self, chunk_id):
        self.chunk_id = chunk_id
        self.readers = 0
        self.writer_active = False
        self.waiting_writers = 0
        self.condition = threading.Condition()

    def acquire_read(self, roll):
        with self.condition:
            while self.writer_active or self.waiting_writers > 0:
                self.condition.wait()
            self.readers += 1
            logger.info(f"[Lock] Roll {roll} acquired READ lock on {self.chunk_id} (readers={self.readers})")

    def release_read(self, roll):
        with self.condition:
            if self.readers <= 0:
                logger.warning(f"[Lock] Roll {roll} attempted to release READ lock on {self.chunk_id} but readers==0")
                return
            self.readers -= 1
            logger.info(f"[Lock] Roll {roll} released READ lock on {self.chunk_id} (readers={self.readers})")
            if self.readers == 0:
                self.condition.notify_all()

    def acquire_write(self, roll):
        with self.condition:
            self.waiting_writers += 1
            try:
                while self.writer_active or self.readers > 0:
                    self.condition.wait()
                self.writer_active = True
                logger.info(f"[Lock] Roll {roll} acquired WRITE lock on {self.chunk_id}")
            finally:
                self.waiting_writers -= 1

    def release_write(self, roll):
        with self.condition:
            if not self.writer_active:
                logger.warning(f"[Lock] Roll {roll} attempted to release WRITE lock on {self.chunk_id} but writer_active==False")
                return
            self.writer_active = False
            logger.info(f"[Lock] Roll {roll} released WRITE lock on {self.chunk_id}")
            self.condition.notify_all()


chunk_locks: Dict[str, ChunkLock] = {}

def init_chunk_locks_from_replication(replication_metadata):
    global chunk_locks
    all_ids = set()
    replicas = replication_metadata.get("replicas", {})
    for replica_id, chunks in replicas.items():
        for chunk_id in chunks.keys():
            all_ids.add(f"{replica_id}:{chunk_id}")
            all_ids.add(chunk_id)
    for cid in all_ids:
        if cid not in chunk_locks:
            chunk_locks[cid] = ChunkLock(cid)
    logger.info(f"[LockManager] Initialized {len(all_ids)} chunk locks")


def _get_chunk_for_roll(roll: str) -> str:
    for chunk_id, rolls in DEFAULT_CHUNK_MAP.items():
        if str(roll) in rolls:
            return chunk_id
    return None


def request_read(roll: str):
    roll = str(roll)
    chunk = _get_chunk_for_roll(roll)
    if not chunk:
        return f"No chunk found for roll {roll}"

    cids = _get_replica_chunks(chunk)
    acquired = []
    try:
        for cid in cids:
            if cid not in chunk_locks:
                chunk_locks[cid] = ChunkLock(cid)
            chunk_locks[cid].acquire_read(roll)
            acquired.append(cid)

        replica_files = replication_metadata.get("replicas", {}).get("replica_1", {})
        path = replica_files.get(chunk, {}).get("path")
        marks = None
        if path:
            try:
                wb = load_workbook(path)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if str(row[0]) == roll:
                        marks = row
                        break
            except Exception as e:
                logger.error(f"[Server] Error reading chunk {chunk}: {e}")
        return marks

    except Exception as e:
        logger.exception(f"[Lock] Exception while acquiring read locks for roll {roll}: {e}")
        for cid in reversed(acquired):
            try:
                chunk_locks[cid].release_read(roll)
            except Exception:
                logger.exception(f"[Lock] Error releasing read lock {cid} during cleanup")
        raise


def release_read(roll: str):
    roll = str(roll)
    chunk = _get_chunk_for_roll(roll)
    if not chunk:
        return False
    for cid in _get_replica_chunks(chunk):
        if cid not in chunk_locks:
            logger.warning(f"[Lock] release_read: no lock object for {cid}")
            continue
        chunk_locks[cid].release_read(roll)
    return True


def request_write(roll: str):
    """
    Acquire write lock for the chunk containing this roll.
    Does NOT update anything, just locks.
    """
    roll = str(roll)
    chunk = _get_chunk_for_roll(roll)
    if not chunk:
        return f"No chunk found for roll {roll}"

    cids = _get_replica_chunks(chunk)
    for cid in cids:
        if cid not in chunk_locks:
            chunk_locks[cid] = ChunkLock(cid)
        chunk_locks[cid].acquire_write(roll)

    logger.info(f"[Server] Write lock acquired for roll={roll}, chunk={chunk}")
    return f"Write lock granted for roll {roll}"


def update_chunk_marks(roll: str, new_marks: int):
    """
    Update ISA marks for a student. Assumes the caller already holds the lock.
    """
    roll = str(roll)
    chunk = _get_chunk_for_roll(roll)
    if not chunk:
        return f"No chunk found for roll {roll}"

    updated = False
    try:
        # Update master Excel
        if not excel_path.exists():
            wb = Workbook()
            ws = wb.active
            ws.append(["Roll", "Name", "Marks", "MCQ", "ISA"])
            wb.save(excel_path)

        wb = load_workbook(excel_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if str(row[0].value) == roll:
                ws.cell(row=row[0].row, column=5, value=new_marks)
                updated = True
                break
        wb.save(excel_path)
        logger.info(f"[Server] Updated MASTER roll={roll} ISA={new_marks}")
    except Exception as e:
        logger.error(f"[Server] Error updating master results.xlsx: {e}")

    # Update replicas
    replicas = replication_metadata.get("replicas", {})
    for replica_id, chunks in replicas.items():
        if chunk in chunks:
            path = Path(chunks[chunk]["path"])
            try:
                wb = load_workbook(path)
                ws = wb.active
                for row in ws.iter_rows(min_row=2):
                    if str(row[0].value) == roll:
                        ws.cell(row=row[0].row, column=5, value=new_marks)
                        updated = True
                        break
                wb.save(path)
                logger.info(f"[Server] Updated roll={roll} marks={new_marks} in {path.name}")
            except Exception as e:
                logger.error(f"[Server] Error updating {path}: {e}")

    return f"Roll {roll} marks updated to {new_marks}" if updated else f"Roll {roll} not found"


def release_write(roll: str):
    """
    Release the write lock for the student’s chunk.
    """
    roll = str(roll)
    chunk = _get_chunk_for_roll(roll)
    if not chunk:
        return False
    for cid in _get_replica_chunks(chunk):
        if cid not in chunk_locks:
            logger.warning(f"[Lock] release_write: no lock object for {cid}")
            continue
        try:
            chunk_locks[cid].release_write(roll)
        except Exception as e:
            logger.error(f"[Lock] Error releasing lock for {cid}: {e}")
    logger.info(f"[Server] Write lock released for roll={roll}, chunk={chunk}")
    return True



def _get_replica_chunks(chunk_id: str) -> list[str]:
    replica_chunks = []
    all_replicas = replication_metadata.get("replicas", {})
    for replica_id, chunks in all_replicas.items():
        if chunk_id in chunks:
            replica_chunks.append(f"{replica_id}:{chunk_id}")
    return sorted(replica_chunks) or [chunk_id]    
                    
# ---------------- Run Server ----------------
def run_server():
    srv = ThreadingXMLRPCServer((SERVER_HOST, SERVER_PORT), allow_none=True, logRequests=False)
    #srv.register_function(cheating_detection, "cheating_detection")
    srv.register_function(input_time, "input_time")
    srv.register_function(get_time, "get_time")
    srv.register_function(receive_cv, "receive_cv")
    srv.register_function(start_synchronization, "start_synchronization")
    srv.register_function(register_student, "register_student")
    srv.register_function(get_registry, "get_registry")
    srv.register_function(register_intent, "register_intent")
    srv.register_function(ok_signal, "ok_signal")
    srv.register_function(update_isa, "update_isa")
    srv.register_function(exam_completed, "exam_completed")
    srv.register_function(accept_backup_result, "accept_backup_result")
    # MCQ endpoints (added)
    srv.register_function(start_mcq, "start_mcq")
    srv.register_function(get_mcq_active, "get_mcq_active")
    srv.register_function(get_question_for_student, "get_question_for_student")
    srv.register_function(submit_mcq_answer, "submit_mcq_answer")
    srv.register_function(submit_mcq_final, "submit_mcq_final")

    srv.register_function(announce_results, "announce_results")
    srv.register_function(request_read, "request_read")
    srv.register_function(release_read, "release_read")
    srv.register_function(request_write, "request_write")
    srv.register_function(release_write, "release_write")
    srv.register_function(update_chunk_marks)



    logger.info(f"[Server] Running on {SERVER_HOST}:{SERVER_PORT} ...")
    # Start server in thread and move control to admin prompt
    threading.Thread(target=srv.serve_forever, daemon=True).start()

    # Admin control loop
    logger.info("[Server] Admin control loop ready. Use commands to drive phases.")
    while True:
        cmd = input(
            "\n[Server Admin] Enter command "
            "(register_students, start_time_sync, start_exam, finish_exam, "
            "start_isa, create_replication, consistency_demo, exit): "
        ).strip()

        if cmd == "register_students":
            logger.info("[Server Admin] Waiting for all students to register...")
            # just an informational pause – students register on their own
            teacher_proxy.phase_complete("Registration")
            client_proxy.phase_complete("Registration")
            for r, url in students_registry.items():
                new_proxy(url).phase_complete("Registration")

        elif cmd == "start_time_sync":
            logger.info("[Server Admin] Triggering Berkeley sync...")
            start_synchronization()
            teacher_proxy.phase_complete("Time Synchronization")
            client_proxy.phase_complete("Time Synchronization")
            for r, url in students_registry.items():
                new_proxy(url).phase_complete("Time Synchronization")

        elif cmd == "start_exam":
            logger.info("[Server Admin] Triggering exam start...")
            try: teacher_proxy.start_exam()
            except Exception: logger.warning("[Server Admin] teacher.start_exam RPC failed.")
            try: client_proxy.start_exam()
            except Exception: logger.warning("[Server Admin] client.start_exam RPC failed.")
            try: start_mcq()
            except Exception as e: logger.warning(f"[Server Admin] WARN starting MCQ: {e}")
            teacher_proxy.phase_complete("Exam Started")
            client_proxy.phase_complete("Exam Started")
            for r, url in students_registry.items():
                new_proxy(url).phase_complete("Exam Started")

        elif cmd == "finish_exam":
            logger.info("[Server Admin] Marking exam as finished and finalizing submissions...")
            exam_completed()
            teacher_proxy.phase_complete("Exam Completed")
            client_proxy.phase_complete("Exam Completed")
            for r, url in students_registry.items():
                new_proxy(url).phase_complete("Exam Completed")

        elif cmd == "start_isa":
            logger.info("[Server Admin] Triggering ISA ask_to_request broadcast...")
            with students_lock:
                for r, url in students_registry.items():
                    try:
                        new_proxy(url).ask_to_request()
                        logger.info(f"[Server Admin] ask_to_request sent to roll {r}")
                    except Exception as e:
                        logger.warning(f"[Server Admin] ask_to_request failed for {r}: {e}")
            teacher_proxy.phase_complete("Ricart-Agrawala ISA")
            client_proxy.phase_complete("Ricart-Agrawala ISA")
            for r, url in students_registry.items():
                new_proxy(url).phase_complete("Ricart-Agrawala ISA")

        elif cmd == "create_replication":
            logger.info("[Server Admin] Starting replication & chunk creation...")
            ok = create_replicas_and_chunks()
            logger.info(f"[Server Admin] Replication {'succeeded' if ok else 'failed'}.")
            teacher_proxy.phase_complete("Replication & Chunking")
            #client_proxy.phase_complete("Replication & Chunking")
            for r, url in students_registry.items():
                new_proxy(url).phase_complete("Replication & Chunking")

        elif cmd == "consistency_demo":
            logger.info("[Server Admin] Consistency demo phase.")
            logger.info("[Server Admin] Only notifying students for Consistency Demo")
            for r, url in students_registry.items():
                try:
                    new_proxy(url).phase_complete("Consistency Demo")
                    new_proxy(url).start_consistency_demo()
                    logger.info(f"[Server Admin] start_consistency_demo sent to roll {r}")
                except Exception as e:
                    logger.warning(f"[Server Admin] consistency_demo RPC failed for {r}: {e}")

        elif cmd == "exit":
            logger.info("[Server Admin] Exiting admin panel.")
            break
        else:
            logger.info("[Server Admin] Unknown command.")
 
    
if __name__ == "__main__":
    run_server()
