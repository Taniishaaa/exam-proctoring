# <---- start of file: updated app.py ---->
from flask import Flask, render_template, request, redirect, url_for
import threading
import time
import random
import datetime
import logging
from pathlib import Path
from openpyxl import Workbook, load_workbook
from flask import flash
from flask import get_flashed_messages
import heapq
import json
import shutil

app = Flask(__name__)
app.secret_key = "supersecretkey123"

# ------------------ GLOBAL STATE ------------------
STUDENTS = {
    "1": {"name": "Swaroop", "marks": 0, "isa": None, "flag": 0, "status": "normal", "cheat_msg": ""},
    "2": {"name": "Tanisha", "marks": 0, "isa": None, "flag": 0, "status": "normal", "cheat_msg": ""},
    "3": {"name": "Siddhesh", "marks": 0, "isa": None, "flag": 0, "status": "normal", "cheat_msg": ""},
    "4": {"name": "Ayush", "marks": 0, "isa": None, "flag": 0, "status": "normal", "cheat_msg": ""},
    "5": {"name": "Nidhi", "marks": 0, "isa": None, "flag": 0, "status": "normal", "cheat_msg": ""},
}

MCQ_QUESTIONS = {
    1: {"question": "Which algorithm is used for clock synchronization?",
        "options": {1: "Lamport", 2: "Berkeley", 3: "Cristian"}, "answer": 2},
    2: {"question": "Which algorithm handles ISA mutual exclusion?",
        "options": {1: "Lamport", 2: "Ricart-Agrawala", 3: "Token Ring"}, "answer": 2},
    3: {"question": "Which library is used for Excel?",
        "options": {1: "pandas", 2: "openpyxl", 3: "xlrd"}, "answer": 2},
    4: {"question": "Exam duration in seconds?",
        "options": {1: "60", 2: "120", 3: "300"}, "answer": 3},
    5: {"question": "Which RPC protocol are we using?",
        "options": {1: "gRPC", 2: "XML-RPC", 3: "RMI"}, "answer": 2},
    6: {"question": "Marks if 1 warning?",
        "options": {1: "100%", 2: "80%", 3: "50%"}, "answer": 2},
    7: {"question": "Marks if 2 warnings?",
        "options": {1: "50%", 2: "80%", 3: "0%"}, "answer": 3},
    8: {"question": "MCQ total marks?",
        "options": {1: "50", 2: "75", 3: "100"}, "answer": 3},
    9: {"question": "Who stores metadata in replication?",
        "options": {1: "Student", 2: "Teacher", 3: "Server"}, "answer": 3},
    10: {"question": "Which data structure used for RA queue?",
         "options": {1: "stack", 2: "heap", 3: "list"}, "answer": 2},
}

EXAM_ACTIVE = False
ISA_PHASE = False
RESULTS_RELEASED = False
EXAM_END_TIME = None
excel_path = Path("results.xlsx")
exam_lock = threading.Lock()

REGISTERED_TEACHER = False
REGISTERED_STUDENTS = set()

TIME_SYNC_PHASE = False
COLLECTED_TIMES = {}
SYNCED_TIMES = {}

# cache live answers for auto-submit
LIVE_ANSWERS = {}

# --- Ricart–Agrawala global state ---
RA_REQUESTS = {}   # roll -> {"ts": int, "requesting": bool, "in_cs": bool}
RA_QUEUE = []      # min-heap [(ts, roll)]
RA_OKS = {}        # roll -> set of OKs received
RA_DEFERRED = {}   # roll -> set of rolls deferred

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(message)s")

from threading import BoundedSemaphore

# --- Backup server simulation ---
MAIN_SERVER_CAPACITY = BoundedSemaphore(3)   # main server handles 3 concurrent requests
MAIN_PROCESSED = 0
BACKUP_PROCESSED = 0
SERVER_LOGS = []

# ------------------ CONSISTENCY DEMO STATE & LOCKS ------------------
CONSISTENCY_PHASE = False
ACTIVE_CONSISTENCY = set()       # rolls participating in demo (set of strings)
CONSISTENCY_HELD = {}            # roll -> "read" / "write" / None (what they currently hold)
replication_metadata = {}        # loaded from replication_metadata.json when available
CHUNK_LOCKS = {}                 # dict of ChunkLock instances keyed by "replica_x:chunkY" and "chunkY"

# ------------------ HELPERS ------------------

def exam_timer():
    """Stops the exam and cheating detection after EXAM_END_TIME and auto-submits using cached answers."""
    global EXAM_ACTIVE
    while EXAM_ACTIVE and EXAM_END_TIME:
        if datetime.datetime.now() >= EXAM_END_TIME:
            EXAM_ACTIVE = False
            logging.info("⌛ Exam ended automatically, stopping cheating detection.")

            def submit_thread(roll):
                """Threaded auto-submit for each student."""
                answers = LIVE_ANSWERS.get(roll, {})
                try:
                    score, server_used = process_submission(roll, answers)
                    logging.info(f"⌛ Auto-submitted Student {roll} with score {score} via {server_used.upper()} server (status={STUDENTS[roll]['status']})")
                except Exception as e:
                    logging.exception(f"Error auto-submitting roll {roll}: {e}")

            threads = []
            for roll in STUDENTS:
                if STUDENTS[roll]["marks"] == 0:
                    t = threading.Thread(target=submit_thread, args=(roll,))
                    t.start()
                    threads.append(t)

            # Wait for all threads to complete before exiting
            for t in threads:
                t.join()

            break
        time.sleep(1)


def grade_mcq(answers: dict) -> int:
    score = 0
    for qid, given in answers.items():
        qid = int(qid)
        if qid in MCQ_QUESTIONS and int(given) == MCQ_QUESTIONS[qid]["answer"]:
            score += 10
    return score

def update_excel(roll, marks, isa=None):
    with exam_lock:
        if not excel_path.exists():
            wb = Workbook()
            ws = wb.active
            ws.append(["Roll", "Name", "Marks/MCQ", "ISA"])
            for r, info in STUDENTS.items():
                ws.append([r, info["name"], info["marks"], info["isa"]])
            wb.save(excel_path)

        wb = load_workbook(excel_path)
        ws = wb.active
        updated = False
        for row in ws.iter_rows(min_row=2):
            if str(row[0].value) == str(roll):
                ws.cell(row=row[0].row, column=3, value=marks)
                if isa is not None:
                    ws.cell(row=row[0].row, column=4, value=isa)
                updated = True
                break
        if not updated:
            ws.append([roll, STUDENTS[roll]["name"], marks, isa])
        wb.save(excel_path)


DEFAULT_CHUNK_MAP = {
    "chunk1": ["1", "2", "3"],
    "chunk2": ["4", "5"]
}
DEFAULT_REPLICATION_FACTOR = 3
METADATA_PATH = Path("replication_metadata.json")

def _read_results_rows():
    if not excel_path.exists():
        logging.error("results.xlsx not found")
        return None, []
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        header = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        rows = [tuple(r) for r in ws.iter_rows(min_row=2, values_only=True)]
        return header, rows
    except Exception as e:
        logging.error(f"Error reading results.xlsx: {e}")
        return None, []

def _filter_rows_for_rolls(rows, rolls):
    rolls_set = set(str(r) for r in rolls)
    return [r for r in rows if str(r[0]) in rolls_set]

def _write_chunk_excel(filepath, header, rows):
    try:
        wb = Workbook()
        ws = wb.active
        ws.append(header)
        for r in rows:
            ws.append(list(r))
        wb.save(filepath)
        logging.info(f"✅ Wrote chunk file: {filepath}")
        return True
    except Exception as e:
        logging.error(f"Failed writing chunk file {filepath}: {e}")
        return False

def create_replicas_and_chunks(replication_factor=DEFAULT_REPLICATION_FACTOR, chunk_map=DEFAULT_CHUNK_MAP):
    header, rows = _read_results_rows()
    if header is None:
        flash("❌ results.xlsx not found or unreadable.", "backup")
        return False

    meta = {
        "created_at": datetime.datetime.now().isoformat(),
        "replication_factor": replication_factor,
        "chunks": {},
        "replicas": {}
    }

    # Prepare chunks
    chunk_contents = {}
    for chunk_id, rolls in chunk_map.items():
        selected = _filter_rows_for_rolls(rows, rolls)
        chunk_contents[chunk_id] = selected
        meta["chunks"][chunk_id] = {"rolls": rolls, "count": len(selected)}

    # Create replicas
    for r_idx in range(1, replication_factor + 1):
        replica_key = f"replica_{r_idx}"
        meta["replicas"][replica_key] = {}
        for chunk_id, content_rows in chunk_contents.items():
            filename = f"replica_{r_idx}_{chunk_id}.xlsx"
            path = Path(filename).resolve()
            ok = _write_chunk_excel(path, header, content_rows)
            meta["replicas"][replica_key][chunk_id] = {
                "path": str(path),
                "rows": len(content_rows)
            }

    # Save metadata
    with open(METADATA_PATH, "w", encoding="utf-8") as fh:
        json.dump(meta, fh, indent=2)
    flash("✅ Replicas and chunks created successfully!", "main")
    logging.info(f"Replication metadata written to {METADATA_PATH}")

    # init chunk locks now that metadata exists
    try:
        init_chunk_locks()
    except Exception:
        logging.exception("Failed to initialize chunk locks after replica creation")

    return True


def process_submission(roll, answers):
    global MAIN_PROCESSED, BACKUP_PROCESSED, SERVER_LOGS

    # --- Try to use Main Server first ---
    acquired = MAIN_SERVER_CAPACITY.acquire(blocking=False)
    if acquired:
        server_used = "main"
        with exam_lock:
            MAIN_PROCESSED += 1
        log_msg = f"✅ Main server processed Student {roll}'s submission."
    else:
        server_used = "backup"
        with exam_lock:
            BACKUP_PROCESSED += 1
        log_msg = f"⚠️ Backup server handled Student {roll}'s submission!"

    # --- Perform grading ---
    score = grade_mcq(answers)
    if STUDENTS[roll]["status"] == "warning":
        score = int(score * 0.5)
    elif STUDENTS[roll]["status"] == "terminated":
        score = 0

    STUDENTS[roll]["marks"] = score
    update_excel(roll, score)

    # --- Record logs ---
    SERVER_LOGS.append(log_msg)

    # Flash only in manual request context
    try:
        from flask import has_request_context
        if has_request_context():
            flash(log_msg, server_used)
    except RuntimeError:
        pass

    # --- Release main server slot if used ---
    if acquired:
        MAIN_SERVER_CAPACITY.release()

    return score, server_used



def simulate_cheating():
    """Randomly issues cheating warnings/terminations while exam is active"""
    global STUDENTS, EXAM_ACTIVE
    while EXAM_ACTIVE:
        roll = random.choice(list(STUDENTS.keys()))
        STUDENTS[roll]["flag"] += 1
        if STUDENTS[roll]["flag"] == 1:
            STUDENTS[roll]["status"] = "warning"
            STUDENTS[roll]["cheat_msg"] = "⚠️ Warning: Cheating detected. Marks will be reduced to 50%."
            logging.info(f"⚠️ Student {roll} caught cheating (1st warning)")
        elif STUDENTS[roll]["flag"] >= 2:
            STUDENTS[roll]["status"] = "terminated"
            STUDENTS[roll]["cheat_msg"] = "⛔ Terminated for repeated cheating. Marks = 0."
            logging.info(f"⛔ Student {roll} terminated for repeated cheating")
        time.sleep(15)

def run_berkeley_sync():
    global SYNCED_TIMES, COLLECTED_TIMES, TIME_SYNC_PHASE
    times = {r: datetime.datetime.strptime(t, "%H:%M:%S") for r, t in COLLECTED_TIMES.items()}
    server_time = times["admin"]

    diffs = {r: (t - server_time).total_seconds() for r, t in times.items()}
    avg_offset = sum(diffs.values()) / len(diffs)

    SYNCED_TIMES = {}
    for r, t in times.items():
        adjusted = t - datetime.timedelta(seconds=(diffs[r] - avg_offset))
        SYNCED_TIMES[r] = adjusted.strftime("%H:%M:%S")

    TIME_SYNC_PHASE = False
    logging.info(f"✅ Berkeley Sync Completed. Synced Times: {SYNCED_TIMES}")

# ------------------ CONSISTENCY: ChunkLock + helpers ------------------

class ChunkLock:
    """Readers-writers lock with writer-preference for a single chunk id."""
    def __init__(self, chunk_id):
        self.chunk_id = chunk_id
        self.readers = 0
        self.writer_active = False
        self.waiting_writers = 0
        self.condition = threading.Condition()

    def acquire_read(self, roll=None):
        with self.condition:
            while self.writer_active or self.waiting_writers > 0:
                self.condition.wait()
            self.readers += 1
            logging.info(f"[Lock] Roll {roll} acquired READ lock on {self.chunk_id} (readers={self.readers})")

    def release_read(self, roll=None):
        with self.condition:
            if self.readers <= 0:
                logging.warning(f"[Lock] Roll {roll} attempted to release READ lock on {self.chunk_id} but readers==0")
                return
            self.readers -= 1
            logging.info(f"[Lock] Roll {roll} released READ lock on {self.chunk_id} (readers={self.readers})")
            if self.readers == 0:
                self.condition.notify_all()

    def acquire_write(self, roll=None):
        with self.condition:
            self.waiting_writers += 1
            try:
                while self.writer_active or self.readers > 0:
                    self.condition.wait()
                self.writer_active = True
                logging.info(f"[Lock] Roll {roll} acquired WRITE lock on {self.chunk_id}")
            finally:
                self.waiting_writers -= 1

    def release_write(self, roll=None):
        with self.condition:
            if not self.writer_active:
                logging.warning(f"[Lock] Roll {roll} attempted to release WRITE lock on {self.chunk_id} but writer_active==False")
                return
            self.writer_active = False
            logging.info(f"[Lock] Roll {roll} released WRITE lock on {self.chunk_id}")
            self.condition.notify_all()


def init_chunk_locks():
    """Initialize CHUNK_LOCKS from replication_metadata.json if present."""
    global replication_metadata, CHUNK_LOCKS
    if not METADATA_PATH.exists():
        logging.info("No replication metadata found to initialize chunk locks.")
        return

    try:
        with open(METADATA_PATH, "r", encoding="utf-8") as fh:
            replication_metadata = json.load(fh)
    except Exception as e:
        logging.error(f"Failed to load replication metadata: {e}")
        replication_metadata = {}
        return

    CHUNK_LOCKS = {}
    # create locks for replica:chunk and chunk
    for replica_id, chunks in replication_metadata.get("replicas", {}).items():
        for chunk_id in chunks.keys():
            CHUNK_LOCKS[f"{replica_id}:{chunk_id}"] = ChunkLock(f"{replica_id}:{chunk_id}")
            # also ensure a lock for the logical chunk id
            if chunk_id not in CHUNK_LOCKS:
                CHUNK_LOCKS[chunk_id] = ChunkLock(chunk_id)

    logging.info(f"[LockManager] Initialized {len(CHUNK_LOCKS)} chunk locks from replication metadata.")


def get_chunk_for_roll(roll):
    """Return chunk id for a given roll using DEFAULT_CHUNK_MAP (mirror of terminal server)."""
    for chunk_id, rolls in DEFAULT_CHUNK_MAP.items():
        if str(roll) in rolls:
            return chunk_id
    return None

def _get_replica_ids_for_chunk(chunk_id):
    """Return list of replica ids (like 'replica_1') that have this chunk from replication_metadata (if available)."""
    reps = []
    try:
        if METADATA_PATH.exists():
            with open(METADATA_PATH, "r", encoding="utf-8") as fh:
                meta = json.load(fh)
            for replica_id, chunks in meta.get("replicas", {}).items():
                if chunk_id in chunks:
                    reps.append(replica_id)
    except Exception as e:
        logging.exception(f"Error reading replication metadata for replicas of {chunk_id}: {e}")
    # fallback if metadata missing: create synthetic replica ids to match default factor
    if not reps:
        for i in range(1, DEFAULT_REPLICATION_FACTOR + 1):
            reps.append(f"replica_{i}")
    return sorted(reps)

def _sorted_lock_keys_for_chunk(chunk_id):
    """Return stable sorted list of CHUNK_LOCKS keys to acquire for a chunk (replica:chunk entries)."""
    reps = _get_replica_ids_for_chunk(chunk_id)
    keys = [f"{rep}:{chunk_id}" for rep in reps]
    # also include the logical chunk id as a lock key (keeps compatibility)
    keys.append(chunk_id)
    # return deterministic order
    return sorted(keys)

def acquire_read_lock(chunk_id, roll):
    """Acquire read lock on all replica locks for chunk_id (in deterministic order)."""
    keys = _sorted_lock_keys_for_chunk(chunk_id)
    for k in keys:
        if k not in CHUNK_LOCKS:
            CHUNK_LOCKS[k] = ChunkLock(k)
        CHUNK_LOCKS[k].acquire_read(roll)

def release_read_lock(chunk_id, roll):
    keys = _sorted_lock_keys_for_chunk(chunk_id)
    for k in keys:
        if k in CHUNK_LOCKS:
            CHUNK_LOCKS[k].release_read(roll)

def acquire_write_lock(chunk_id, roll):
    """Acquire write lock on all replica locks for chunk_id (in deterministic order)."""
    keys = _sorted_lock_keys_for_chunk(chunk_id)
    for k in keys:
        if k not in CHUNK_LOCKS:
            CHUNK_LOCKS[k] = ChunkLock(k)
        CHUNK_LOCKS[k].acquire_write(roll)

def release_write_lock(chunk_id, roll):
    keys = _sorted_lock_keys_for_chunk(chunk_id)
    for k in keys:
        if k in CHUNK_LOCKS:
            CHUNK_LOCKS[k].release_write(roll)

def try_acquire_write_lock(chunk_id, roll):
    """
    Attempt to acquire a write lock immediately (non-blocking).
    Returns (True, msg) if successful, else (False, reason).
    """
    keys = _sorted_lock_keys_for_chunk(chunk_id)
    for k in keys:
        if k not in CHUNK_LOCKS:
            CHUNK_LOCKS[k] = ChunkLock(k)
        lock = CHUNK_LOCKS[k]
        with lock.condition:
            if lock.writer_active:
                return False, "being written by another student"
            if lock.readers > 0:
                return False, "being read by other students"
        # No active readers/writers → acquire
        lock.writer_active = True
        logging.info(f"[Lock] Roll {roll} non-blocking acquired WRITE lock on {k}")
    return True, "acquired"


def get_marks_from_results(roll):
    """Return ISA marks (or full row) for roll from results.xlsx"""
    header, rows = _read_results_rows()
    if header is None:
        return None
    for r in rows:
        if str(r[0]) == str(roll):
            # r may be (roll, name, marks, mcq, isa) or similar; we return row as tuple
            return r
    return None

def update_chunk_marks_for_chunk_and_replicas(roll, new_marks):
    """
    Update master results.xlsx and then update all replica chunk files for the chunk that contains roll.
    """
    roll = str(roll)
    chunk = get_chunk_for_roll(roll)
    if not chunk:
        logging.error(f"No chunk found for roll {roll} during update_chunk_marks.")
        return False

    # Update master
    try:
        with exam_lock:
            if not excel_path.exists():
                wb = Workbook()
                ws = wb.active
                ws.append(["Roll", "Name", "Marks/MCQ", "ISA"])
                wb.save(excel_path)

            wb = load_workbook(excel_path)
            ws = wb.active
            updated = False
            for row in ws.iter_rows(min_row=2):
                if str(row[0].value) == roll:
                    # find ISA column index; assume last column is ISA if present
                    isa_col = 5 if ws.max_column >= 5 else ws.max_column
                    # ensure we have 5 columns
                    while ws.max_column < 5:
                        ws.cell(row=1, column=ws.max_column+1, value=None)
                    ws.cell(row=row[0].row, column=4, value=int(new_marks))
                    updated = True
                    break
            if not updated:
                ws.append([roll, STUDENTS.get(roll, {}).get("name", f"Student{roll}"), "NA", "NA", int(new_marks)])
            wb.save(excel_path)
    except Exception as e:
        logging.exception(f"Error updating master excel for roll {roll}: {e}")
        return False

    # Read master rows to propagate
    header, rows = _read_results_rows()
    if header is None:
        logging.error("Master file disappeared after update.")
        return False

    # Update each replica file for the chunk
    try:
        if METADATA_PATH.exists():
            with open(METADATA_PATH, "r", encoding="utf-8") as fh:
                meta = json.load(fh)
        else:
            meta = {}
        for replica_id, chunks in meta.get("replicas", {}).items():
            if chunk in chunks:
                path = Path(chunks[chunk]["path"])
                # filter rows for this chunk's rolls
                rolls_list = DEFAULT_CHUNK_MAP.get(chunk, [])
                selected = _filter_rows_for_rolls(rows, rolls_list)
                _write_chunk_excel(path, header, selected)
    except Exception as e:
        logging.exception(f"Error updating replicas for chunk {chunk}: {e}")
        # even on failure we consider master updated; propagate attempts logged
    return True

# ------------------ ROUTES ------------------
@app.route("/")
def home():
    return render_template("home.html")

@app.route("/register_teacher", methods=["POST"])
def register_teacher():
    global REGISTERED_TEACHER
    if REGISTERED_TEACHER:
        return redirect(url_for("teacher_panel"))
    REGISTERED_TEACHER = True
    return redirect(url_for("teacher_panel"))

@app.route("/register_student", methods=["POST"])
def register_student():
    global REGISTERED_STUDENTS
    available = [r for r in STUDENTS.keys() if r not in REGISTERED_STUDENTS]
    if not available:
        return "All students already registered.", 400
    roll = available[0]
    REGISTERED_STUDENTS.add(roll)
    return redirect(url_for("student_portal", roll=roll))


@app.route("/admin")
def admin_panel():
    messages = get_flashed_messages(with_categories=True)
    return render_template(
        "admin.html",
        exam_active=EXAM_ACTIVE,
        isa_phase=ISA_PHASE,
        results_released=RESULTS_RELEASED,
        time_sync=TIME_SYNC_PHASE,
        synced_times=SYNCED_TIMES,
        main_processed=MAIN_PROCESSED,
        backup_processed=BACKUP_PROCESSED,
        server_logs=SERVER_LOGS,
        messages=messages,
        replication_done=REPLICATION_DONE,
        consistency_phase=CONSISTENCY_PHASE,
    )

@app.route("/admin/start_exam", methods=["POST"])
def start_exam():
    global EXAM_ACTIVE, EXAM_END_TIME
    EXAM_ACTIVE = True
    EXAM_END_TIME = datetime.datetime.now() + datetime.timedelta(seconds=30)
    logging.info("🚀 Exam started for 30s")
    threading.Thread(target=simulate_cheating, daemon=True).start()
    threading.Thread(target=exam_timer, daemon=True).start()
    return redirect(url_for("admin_panel"))

@app.route("/admin/start_sync", methods=["POST"])
def start_sync():
    global TIME_SYNC_PHASE, COLLECTED_TIMES
    TIME_SYNC_PHASE = True
    COLLECTED_TIMES = {}
    logging.info("🚀 Time Sync Phase started")
    return redirect(url_for("admin_sync"))

@app.route("/admin/sync", methods=["GET", "POST"])
def admin_sync():
    global COLLECTED_TIMES
    if request.method == "POST":
        local_time = request.form["local_time"]
        COLLECTED_TIMES["admin"] = local_time
        logging.info(f"Admin submitted time {local_time}")
        if len(COLLECTED_TIMES) == 7:
            run_berkeley_sync()
        return redirect(url_for("admin_panel"))
    return render_template("sync.html", role="Admin")

@app.route("/exam_status")
def exam_status():
    global EXAM_ACTIVE, EXAM_END_TIME
    remaining = 0
    if EXAM_ACTIVE and EXAM_END_TIME:
        delta = (EXAM_END_TIME - datetime.datetime.now()).total_seconds()
        remaining = max(0, int(delta))
    return {"active": EXAM_ACTIVE, "remaining": remaining}


@app.route("/admin/start_isa", methods=["POST"])
def start_isa():
    global ISA_PHASE
    ISA_PHASE = True
    logging.info("🚀 ISA Phase started")
    return redirect(url_for("admin_panel"))


@app.route("/student/<roll>/isa_request", methods=["POST"])
def isa_request(roll):
    ts = int(time.time() * 1000000)  # microsecond timestamp
    RA_REQUESTS[roll] = {"ts": ts, "requesting": True, "in_cs": False}
    RA_OKS[roll] = set()
    RA_DEFERRED[roll] = set()
    heapq.heappush(RA_QUEUE, (ts, roll))

    logging.info(f"📥 Student {roll} requested ISA at ts={ts}")

    # --- Compare timestamps and decide OK/defer for each peer ---
    for other, state in RA_REQUESTS.items():
        if other == roll:
            continue

        # Peer is idle → immediate OK
        if not state.get("requesting") and not state.get("in_cs"):
            RA_OKS[roll].add(other)
            logging.info(f"✅ Student {other} (idle) gave OK to {roll}")
            continue

        # Peer is in CS → must defer
        if state.get("in_cs"):
            RA_DEFERRED[other].add(roll)
            logging.info(f"⏸ Student {roll} deferred by {other} (in CS)")
            continue

        # Both requesting → compare (timestamp, roll) deterministically
        my_req = (ts, roll)
        other_req = (state["ts"], other)

        if other_req < my_req:
            # Peer requested earlier (or has smaller roll number on tie)
            RA_DEFERRED[other].add(roll)
            logging.info(f"⏸ Student {roll} deferred by {other} (older ts or lower roll)")
        else:
            # Current student has priority → peer should defer
            RA_OKS[roll].add(other)
            logging.info(f"✅ Student {other} gave OK to {roll} (newer ts or higher roll)")

    # --- Check if this student already got all OKs and can enter CS ---
    peers = set(RA_REQUESTS.keys()) - {roll}
    if peers.issubset(RA_OKS[roll]):
        RA_REQUESTS[roll]["in_cs"] = True
        logging.info(f"🚪 Student {roll} enters CS immediately")

    return redirect(url_for("student_check_entry", roll=roll))


@app.route("/admin/release_results", methods=["POST"])
def release_results():
    global RESULTS_RELEASED
    RESULTS_RELEASED = True
    logging.info("✅ Results released")
    return redirect(url_for("admin_panel"))

REPLICATION_DONE = False
@app.route("/admin/create_replica", methods=["POST"])
def create_replica():
    ok = create_replicas_and_chunks()
    if ok:
        REPLICATION_DONE = True
        ISA_PHASE = False 
        logging.info("Replication & chunk creation completed via Admin panel.")
    else:
        logging.error("Replication & chunk creation failed.")
    return redirect(url_for("admin_panel"))


@app.route("/teacher", methods=["GET", "POST"])
def teacher_panel():
    global TIME_SYNC_PHASE
    if not REGISTERED_TEACHER:
        return redirect(url_for("home"))

    if TIME_SYNC_PHASE and "teacher" not in COLLECTED_TIMES:
        if request.method == "POST":
            local_time = request.form["local_time"]
            COLLECTED_TIMES["teacher"] = local_time
            logging.info(f"Teacher submitted time {local_time}")
            if len(COLLECTED_TIMES) == 7:
                run_berkeley_sync()
            return render_template("submitted.html", role="Teacher")
        return render_template("sync.html", role="Teacher")

    if not TIME_SYNC_PHASE and "teacher" in SYNCED_TIMES:
        return render_template("synced.html", role="Teacher", time=SYNCED_TIMES["teacher"])

    return render_template("teacher.html", students=STUDENTS,
                           results_released=RESULTS_RELEASED,
                           synced_times=SYNCED_TIMES if SYNCED_TIMES else None)


@app.route("/student/<roll>", methods=["GET", "POST"])
def student_portal(roll):
    global TIME_SYNC_PHASE, EXAM_ACTIVE, EXAM_END_TIME, ISA_PHASE, CONSISTENCY_PHASE, ACTIVE_CONSISTENCY

    if roll not in STUDENTS:
        return f"Invalid roll number {roll}", 404

    # ---------------- Time Sync Phase ----------------
    if TIME_SYNC_PHASE and roll not in COLLECTED_TIMES:
        if request.method == "POST":
            local_time = request.form["local_time"]
            COLLECTED_TIMES[roll] = local_time
            logging.info(f"Student {roll} submitted time {local_time}")
            if len(COLLECTED_TIMES) == 7:
                run_berkeley_sync()
            return render_template("submitted.html", role=f"Student {roll}")
        return render_template("sync.html", role=f"Student {roll}")

    # --- If student already submitted (manual or auto) and waiting for ISA ---
    if not EXAM_ACTIVE and STUDENTS[roll]["marks"] > 0 and STUDENTS[roll]["isa"] is None and not ISA_PHASE:
        return render_template("student_submitted.html", roll=roll)

    # After sync but before exam
    if not TIME_SYNC_PHASE and roll in SYNCED_TIMES and not EXAM_ACTIVE and not ISA_PHASE:
        return render_template("synced.html", role=f"Student {roll}", time=SYNCED_TIMES[roll])

    # ---------------- Exam Phase ----------------
    if EXAM_ACTIVE:
        # If this student has already submitted, show confirmation instead of restarting exam
        if STUDENTS[roll]["marks"] > 0:
            return render_template("student_submitted.html", roll=roll)
        return redirect(url_for("student_exam", roll=roll, qid=1))

    # ---------------- ISA Phase ----------------
    if ISA_PHASE and STUDENTS[roll]["isa"] is None:
        # Prompt student to join ISA entry (Yes/No)
        return render_template("student_isa_prompt.html", roll=roll)

    # ---------------- Consistency Demo Phase ----------------
    if CONSISTENCY_PHASE:
        # Add student to active consistency set (only once)
        if str(roll) not in ACTIVE_CONSISTENCY:
            ACTIVE_CONSISTENCY.add(str(roll))
            logging.info(f"Student {roll} joined Consistency Demo")
        # Show the consistency prompt page with Read / Write / Exit Demo buttons
        return render_template("consistency_prompt.html", roll=roll)

    # ---------------- Waiting Phase ----------------
    return render_template("student_wait.html", roll=roll)



@app.route("/student/<roll>/exam/<int:qid>", methods=["GET", "POST"])
def student_exam(roll, qid):
    global LIVE_ANSWERS

    if roll not in STUDENTS:
        return f"Invalid roll {roll}", 404

    # If exam is not active, send them back to student_portal (which will render appropriate page)
    if not EXAM_ACTIVE:
        return redirect(url_for("student_portal", roll=roll))

    now = datetime.datetime.now()
    # If exam time already passed, redirect to results (exam_timer will perform auto-submits)
    if EXAM_END_TIME and now > EXAM_END_TIME:
        return redirect(url_for("results", roll=roll))

    # Save answer if POST (also handles Next/Prev/Submit)
    if request.method == "POST":
        ans = request.form.get("answer")
        if ans is not None:
            # store as string; key is qid (int)
            LIVE_ANSWERS.setdefault(roll, {})[qid] = ans

        # Navigation
        if "next" in request.form and qid < len(MCQ_QUESTIONS):
            return redirect(url_for("student_exam", roll=roll, qid=qid+1))
        elif "prev" in request.form and qid > 1:
            return redirect(url_for("student_exam", roll=roll, qid=qid-1))
        elif "submit" in request.form:
            # Final submit -> use process_submission so main/backup logic applies
            answers = LIVE_ANSWERS.get(roll, {})
            score, server_used = process_submission(roll, answers)
            logging.info(f"✅ Student {roll} submitted with score {score} via {server_used.upper()} server (status={STUDENTS[roll]['status']})")
            return redirect(url_for("student_portal", roll=roll))

    # For GET: determine navigation, selected option from LIVE_ANSWERS
    prev_qid = qid - 1 if qid > 1 else None
    next_qid = qid + 1 if qid < len(MCQ_QUESTIONS) else None
    selected = LIVE_ANSWERS.get(roll, {}).get(qid)

    # Remaining time safety (0 if no EXAM_END_TIME)
    remaining = int((EXAM_END_TIME - now).total_seconds()) if EXAM_END_TIME else 0
    if remaining < 0:
        remaining = 0

    # Render single-question template (student.html expects these variables)
    return render_template("student.html",
                           roll=roll,
                           qid=qid,
                           question=MCQ_QUESTIONS[qid],
                           prev_qid=prev_qid,
                           next_qid=next_qid,
                           selected=selected,
                           remaining=remaining,
                           student=STUDENTS[roll])


@app.route("/student/<roll>/results")
def results(roll):
    if not RESULTS_RELEASED:
        return f"Results not released yet for roll {roll}"
    student = STUDENTS.get(roll)
    return render_template("results.html", student=student)


@app.route("/student/<roll>/isa_check")
def student_check_entry(roll):
    global RA_QUEUE

    if roll not in RA_REQUESTS:
        return f"Invalid ISA request for Student {roll}", 400

    peers = set(RA_REQUESTS.keys()) - {roll}
    received_oks = RA_OKS.get(roll, set())

    # --- If all OKs received → enter CS ---
    if peers.issubset(received_oks):
        RA_REQUESTS[roll]["in_cs"] = True
        logging.info(f"🚪 Student {roll} enters CS")
        return render_template("student_isa_entry.html", roll=roll)

    # --- Determine students ahead of current one (true queue order) ---
    sorted_queue = [r for (_, r) in sorted(RA_QUEUE)]

    # Students ahead = all those who appear before this roll in queue & are still requesting
    ahead_in_queue = []
    for r in sorted_queue:
        if r == roll:
            break
        if RA_REQUESTS.get(r, {}).get("requesting", False):
            ahead_in_queue.append(r)

    logging.info(f"Student {roll} is waiting behind {ahead_in_queue}")

    # --- Render waiting state ---
    return f"Student {roll} is waiting behind {ahead_in_queue}"


@app.route("/student/<roll>/isa_submit", methods=["POST"])
def isa_submit(roll):
    global RA_QUEUE

    marks = int(request.form["isa_marks"])
    STUDENTS[roll]["isa"] = marks
    update_excel(roll, STUDENTS[roll]["marks"], isa=marks)

    # Exit CS
    RA_REQUESTS[roll]["in_cs"] = False
    RA_REQUESTS[roll]["requesting"] = False
    logging.info(f"📤 Student {roll} submitted ISA={marks} and exited CS")

    # Flush deferred OKs
    for other in list(RA_DEFERRED[roll]):
        RA_OKS[other].add(roll)
        logging.info(f"➡️ Student {roll} sent deferred OK to {other}")
    RA_DEFERRED[roll].clear()

    # 🧠 Re-evaluate global RA queue (priority = timestamp)
    heapq.heapify(RA_QUEUE)
    active = [r for (_, r) in RA_QUEUE if RA_REQUESTS[r]["requesting"]]

    if active:
        # The earliest (ts, roll) in queue gets the next CS turn
        next_roll = active[0]
        if not RA_REQUESTS[next_roll]["in_cs"]:
            # Give OKs from all idle/non-CS peers
            for peer, state in RA_REQUESTS.items():
                if peer == next_roll:
                    continue
                if not state["in_cs"]:
                    RA_OKS[next_roll].add(peer)

            peers = set(RA_REQUESTS.keys()) - {next_roll}
            if peers.issubset(RA_OKS[next_roll]):
                RA_REQUESTS[next_roll]["in_cs"] = True
                logging.info(f"🚪 Queue-based entry: Student {next_roll} enters CS next")

    # 🔁 Safety sweep — check if any other waiting student now qualifies
    for sid, state in RA_REQUESTS.items():
        if not state["in_cs"] and state["requesting"]:
            peers = set(RA_REQUESTS.keys()) - {sid}
            if peers.issubset(RA_OKS.get(sid, set())):
                RA_REQUESTS[sid]["in_cs"] = True
                logging.info(f"🚪 Student {sid} re-evaluated and now enters CS")

    # ✅ Check if ISA phase completed for everyone
    if all(s.get("isa") is not None for s in STUDENTS.values()):
        logging.info("✅ ISA phase completed for all students.")

    return redirect(url_for("student_portal", roll=roll))



# ------------------ CONSISTENCY ROUTES ------------------

@app.route("/admin/start_consistency", methods=["POST"])
def start_consistency():
    """Admin triggers the consistency demo for all students."""
    global CONSISTENCY_PHASE, ACTIVE_CONSISTENCY, CONSISTENCY_HELD
    CONSISTENCY_PHASE = True
    # add all registered students (or all STUDENTS) to active demo
    ACTIVE_CONSISTENCY = set(str(r) for r in STUDENTS.keys())
    CONSISTENCY_HELD = {}
    logging.info("🧩 Consistency demo started by Admin; prompting all students.")
    flash("🧩 Consistency demo started — students will see the prompt.", "main")
    return redirect(url_for("admin_panel"))


@app.route("/student/<roll>/consistency/read", methods=["GET"])
def consistency_read(roll):
    global CONSISTENCY_PHASE, ACTIVE_CONSISTENCY, CONSISTENCY_HELD

    if not CONSISTENCY_PHASE or str(roll) not in ACTIVE_CONSISTENCY:
        return redirect(url_for("student_portal", roll=roll))

    chunk = get_chunk_for_roll(roll)
    keys = _sorted_lock_keys_for_chunk(chunk)

    # Check for active writer
    for k in keys:
        if k not in CHUNK_LOCKS:
            CHUNK_LOCKS[k] = ChunkLock(k)
        lock = CHUNK_LOCKS[k]
        with lock.condition:
            if lock.writer_active:
                logging.info(f"[Lock] Roll {roll} waiting — writer active on {chunk}")
                return render_template(
                    "consistency_wait.html",
                    roll=roll,
                    lock_type="WRITE",
                    retry_url=url_for("consistency_read", roll=roll)
                )

    # Acquire read locks
    acquire_read_lock(chunk, roll)
    CONSISTENCY_HELD[roll] = "read"
    logging.info(f"[Consistency] Student {roll} acquired READ lock on {chunk}")

    # Fetch marks
    marks = "N/A"
    try:
        wb = load_workbook("results.xlsx")
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == str(roll):
                marks = row[3] if len(row) > 3 else "N/A"
                break
        wb.close()
    except Exception as e:
        logging.error(f"[Error] Could not fetch marks for {roll}: {e}")

    return render_template("consistency_read.html", roll=roll, marks=marks, chunk=chunk)


@app.route("/student/<roll>/consistency/write", methods=["GET", "POST"])
def consistency_write(roll):
    global CONSISTENCY_PHASE, ACTIVE_CONSISTENCY, CONSISTENCY_HELD

    if not CONSISTENCY_PHASE or str(roll) not in ACTIVE_CONSISTENCY:
        return redirect(url_for("student_portal", roll=roll))

    chunk = get_chunk_for_roll(roll)

    # ---------- GET: acquire lock + show marks ----------
    if request.method == "GET":
        ok, reason = try_acquire_write_lock(chunk, roll)
        if not ok:
            logging.info(f"[Lock] Roll {roll} waiting: {reason} on {chunk}")
            return render_template(
                "consistency_wait.html",
                roll=roll,
                lock_type=reason.upper(),
                retry_url=url_for("consistency_write", roll=roll)
            )

        CONSISTENCY_HELD[roll] = "write"

        marks = "N/A"
        try:
            wb = load_workbook("results.xlsx")
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if str(row[0]) == str(roll):
                    marks = row[3] if len(row) > 3 else "N/A"
                    break
            wb.close()
        except Exception as e:
            logging.error(f"[Error] Fetching marks for write {roll}: {e}")

        return render_template("consistency_write.html", roll=roll, marks=marks, chunk=chunk)

    # ---------- POST: update marks & release lock ----------
    new_marks = int(request.form["isa_marks"])
    try:
        update_chunk_marks_for_chunk_and_replicas(roll, new_marks)
        flash("✅ Marks updated and replicated successfully.", "success")
        logging.info(f"[Consistency] Roll {roll} updated marks={new_marks} in {chunk}")
    except Exception as e:
        logging.exception(f"[Error] Updating marks for roll {roll}: {e}")
        flash("❌ Error updating marks.", "error")
    finally:
        release_write_lock(chunk, roll)
        CONSISTENCY_HELD.pop(roll, None)
        logging.info(f"[Lock] Roll {roll} released WRITE lock on {chunk}")

    return render_template("consistency_prompt.html", roll=roll)



@app.route("/student/<roll>/consistency/exit_cs", methods=["POST"])
def consistency_exit_cs(roll):
    roll = str(roll)
    held = CONSISTENCY_HELD.get(roll)
    chunk = get_chunk_for_roll(roll)

    if held == "read":
        release_read_lock(chunk, roll)
        logging.info(f"[Consistency] Student {roll} released READ locks on {chunk}")
    elif held == "write":
        release_write_lock(chunk, roll)
        logging.info(f"[Consistency] Student {roll} released WRITE locks on {chunk}")

    CONSISTENCY_HELD.pop(roll, None)
    flash("🔓 Lock released successfully.", "main")
    return redirect(url_for("student_portal", roll=roll))


@app.route("/student/<roll>/consistency/exit_demo")
def consistency_exit_demo(roll):
    """Remove student from the active consistency demo and return to waiting screen."""
    roll = str(roll)
    ACTIVE_CONSISTENCY.discard(roll)
    CONSISTENCY_HELD.pop(roll, None)
    logging.info(f"[Consistency] Student {roll} exited the demo.")
    return render_template("student_wait.html", roll=roll)


@app.route("/status")
def status():
    return {
        roll: {
            "name": info["name"],
            "status": info["status"],
            "marks": info["marks"],
            "flags": info["flag"],
            "cheat_msg": info["cheat_msg"],
        }
        for roll, info in STUDENTS.items()
    }

if __name__ == "__main__":
    # Try to initialize chunk locks at startup (if metadata is present)
    try:
        init_chunk_locks()
    except Exception:
        logging.info("No replication metadata at startup or failed to init locks.")
    app.run(debug=True)
# <---- end of file: updated app.py ---->
